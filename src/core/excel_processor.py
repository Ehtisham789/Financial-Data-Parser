
import pandas as pd
import openpyxl

class ExcelProcessor:
    def __init__(self):
        self.files = {}

    def load_files(self, file_paths):
        for file_path in file_paths:
            try:
                # Try to read with pandas first
                xls = pd.ExcelFile(file_path)
                self.files[file_path] = {
                    'pandas_excel_file': xls,
                    'sheets': {sheet_name: pd.read_excel(xls, sheet_name=sheet_name) for sheet_name in xls.sheet_names}
                }
                print(f"Successfully loaded {file_path} with pandas.")
            except Exception as e:
                print(f"Could not load {file_path} with pandas: {e}")
                try:
                    # Fallback to openpyxl if pandas fails
                    workbook = openpyxl.load_workbook(file_path)
                    self.files[file_path] = {
                        'openpyxl_workbook': workbook,
                        'sheets': {sheet_name: None for sheet_name in workbook.sheetnames} # Data will be extracted on demand
                    }
                    print(f"Successfully loaded {file_path} with openpyxl.")
                except Exception as e_openpyxl:
                    print(f"Could not load {file_path} with openpyxl: {e_openpyxl}")

    def get_sheet_info(self):
        info = {}
        for file_path, file_data in self.files.items():
            file_info = {'sheets': {}}
            if 'pandas_excel_file' in file_data:
                for sheet_name, df in file_data['sheets'].items():
                    file_info['sheets'][sheet_name] = {
                        'dimensions': df.shape,
                        'column_names': df.columns.tolist()
                    }
            elif 'openpyxl_workbook' in file_data:
                workbook = file_data['openpyxl_workbook']
                for sheet_name in workbook.sheetnames:
                    # For openpyxl, we can't get dimensions/column names without loading the sheet
                    # This is a placeholder, actual data extraction would be in extract_data
                    file_info['sheets'][sheet_name] = {
                        'dimensions': 'N/A (load with extract_data)',
                        'column_names': 'N/A (load with extract_data)'
                    }
            info[file_path] = file_info
        return info

    def extract_data(self, file_path, sheet_name):
        if file_path in self.files:
            file_data = self.files[file_path]
            if 'pandas_excel_file' in file_data:
                return file_data['sheets'].get(sheet_name)
            elif 'openpyxl_workbook' in file_data:
                workbook = file_data['openpyxl_workbook']
                if sheet_name in workbook.sheetnames:
                    # Load sheet data into a pandas DataFrame
                    sheet = workbook[sheet_name]
                    data = sheet.values
                    cols = next(data) # Get header row
                    df = pd.DataFrame(data, columns=cols)
                    # Store the DataFrame for future access
                    self.files[file_path]['sheets'][sheet_name] = df
                    return df
        return None

    def preview_data(self, file_path, sheet_name, rows=5):
        df = self.extract_data(file_path, sheet_name)
        if df is not None:
            print(f"\nPreview of {sheet_name} from {file_path}:")
            print(df.head(rows))
        else:
            print(f"Could not preview data for {sheet_name} from {file_path}.")


